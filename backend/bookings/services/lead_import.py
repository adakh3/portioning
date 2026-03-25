import csv
import io
import re
from dataclasses import dataclass, field
from datetime import datetime

from bookings.models import Lead


def parse_event_date(raw):
    """Best-effort parse of freeform date strings like '28 march', 'April 11', '10 june 2026'."""
    if not raw or "test lead" in str(raw).lower() or "not confirm" in str(raw).lower():
        return None

    raw = str(raw).strip()

    for fmt in ("%d %B %Y", "%d %B", "%B %d", "%d %b %Y", "%d %b", "%b %d",
                "%d %B, %Y", "%B %Y", "%B"):
        try:
            dt = datetime.strptime(raw, fmt)
            if dt.year == 1900:
                dt = dt.replace(year=2026)
            return dt.date()
        except ValueError:
            continue

    match = re.search(
        r"(\d{1,2})\s*(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s*(\d{4})?",
        raw, re.I,
    )
    if match:
        day, month_str, year_str = match.groups()
        year = int(year_str) if year_str else 2026
        try:
            return datetime.strptime(f"{day} {month_str} {year}", "%d %b %Y").date()
        except ValueError:
            pass

    match = re.search(
        r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\w*\s*(\d{1,2})?,?\s*(\d{4})?",
        raw, re.I,
    )
    if match:
        month_str, day_str, year_str = match.groups()
        day = int(day_str) if day_str else 1
        year = int(year_str) if year_str else 2026
        try:
            return datetime.strptime(f"{day} {month_str} {year}", "%d %b %Y").date()
        except ValueError:
            pass

    return None


@dataclass
class ImportRow:
    row_num: int = 0
    contact_name: str = ""
    contact_email: str = ""
    contact_phone: str = ""
    event_type: str = ""
    guest_estimate: int | None = None
    guest_estimate_raw: str = ""
    event_date: object = None  # date or None
    lead_date: object = None  # date or None
    source: str = ""
    status: str = ""
    notes: str = ""
    product_name: str = ""
    assigned_to_email: str = ""
    skipped: bool = False
    skip_reason: str = ""
    error: str = ""
    created: bool = False
    duplicate_warning: bool = False


def parse_rows(data_rows, header):
    """Parse raw spreadsheet rows into ImportRow objects. Pure structural parser, no DB access."""
    col = {name: i for i, name in enumerate(header)}
    required = ["full_name", "phone_number", "event_type"]
    for r in required:
        if r not in col:
            raise ValueError(f"Missing required column: {r}")

    results = []
    for i, row in enumerate(data_rows, start=2):
        ir = ImportRow(row_num=i)

        name = str(row[col["full_name"]] or "").strip()
        phone = str(row[col["phone_number"]] or "").strip()
        email = str(row[col["email"]] or "").strip() if "email" in col else ""

        # Strip p: prefix from phone numbers
        phone = re.sub(r"^p:", "", phone).strip()

        # Skip test leads
        if "test lead" in name.lower() or "test lead" in email.lower():
            ir.contact_name = name
            ir.contact_email = email
            ir.skipped = True
            ir.skip_reason = "Test lead"
            results.append(ir)
            continue

        # Skip empty rows (name and phone both empty)
        if not name and not phone:
            ir.skipped = True
            ir.skip_reason = "Empty row"
            results.append(ir)
            continue

        # Collect all errors for this row
        errors = []
        if not name:
            errors.append("Missing full_name")
        if not phone:
            errors.append("Missing phone_number")

        event_type_raw = str(row[col.get("event_type", "")] or "").strip().lower()

        # Parse guest count as plain integer
        guest_raw = str(row[col.get("your_guests", "")] or "").strip()
        guest_estimate = None
        if guest_raw:
            digits = re.sub(r"[^\d]", "", guest_raw)
            if digits:
                guest_estimate = int(digits)
            # If non-empty but not parseable as int, validation step will flag it

        date_raw = row[col.get("your_event_date", "")]
        event_date = parse_event_date(date_raw)

        lead_date_raw = row[col.get("lead_date", "")]
        lead_date = parse_event_date(lead_date_raw)
        # Also try datetime objects (common in xlsx)
        if not lead_date and lead_date_raw:
            if hasattr(lead_date_raw, 'date'):
                lead_date = lead_date_raw.date()
            elif hasattr(lead_date_raw, 'year'):
                lead_date = lead_date_raw

        platform = str(row[col.get("platform", "")] or "").strip().lower()
        status_raw = str(row[col.get("lead_status", "")] or "").strip().lower()

        campaign = str(row[col.get("campaign_name", "")] or "").strip()
        notes_parts = []
        if campaign:
            notes_parts.append(f"Campaign: {campaign}")
        if date_raw and not event_date:
            notes_parts.append(f"Date (unparsed): {date_raw}")

        product_name = str(row[col["product"]] or "").strip() if "product" in col else ""
        assigned_to_email = str(row[col["assigned_to"]] or "").strip() if "assigned_to" in col else ""

        ir.contact_name = name[:200]
        ir.contact_email = email[:254]
        ir.contact_phone = phone[:50]
        ir.event_type = event_type_raw
        ir.guest_estimate = guest_estimate
        ir.guest_estimate_raw = guest_raw
        ir.event_date = event_date
        ir.lead_date = lead_date
        ir.source = platform
        ir.status = status_raw
        ir.notes = "\n".join(notes_parts)
        ir.product_name = product_name
        ir.assigned_to_email = assigned_to_email

        if errors:
            ir.error = "; ".join(errors)

        results.append(ir)

    return results


def validate_rows(import_rows, org):
    """Validate parsed rows against the org's configured options. Sets row.error for invalid data."""
    from bookings.models.choices import EventTypeOption, SourceOption, LeadStatusOption
    from bookings.models.leads import ProductLine

    # Load org's valid options once
    event_types = set(
        EventTypeOption.objects.filter(organisation=org, is_active=True)
        .values_list('value', flat=True)
    )
    sources = set(
        SourceOption.objects.filter(organisation=org, is_active=True)
        .values_list('value', flat=True)
    )
    statuses = set(
        LeadStatusOption.objects.filter(organisation=org, is_active=True)
        .values_list('value', flat=True)
    )
    products = {
        p.name.lower(): p
        for p in ProductLine.objects.filter(organisation=org, is_active=True)
    }

    for row in import_rows:
        if row.skipped:
            continue

        # Start with any errors already set by parse_rows (missing name/phone)
        errors = row.error.split("; ") if row.error else []

        # event_type: must match org's options
        if row.event_type and row.event_type not in event_types:
            valid = ", ".join(sorted(event_types))
            errors.append(f"Invalid event_type '{row.event_type}' — valid: {valid}")
        elif not row.event_type:
            errors.append("Missing event_type")

        # source: optional, defaults to "website"
        if row.source and row.source not in sources:
            valid = ", ".join(sorted(sources))
            errors.append(f"Invalid platform '{row.source}' — valid: {valid}")
        elif not row.source:
            row.source = "website"

        # status: optional, defaults to "new"
        if row.status and row.status not in statuses:
            valid = ", ".join(sorted(statuses))
            errors.append(f"Invalid lead_status '{row.status}' — valid: {valid}")
        elif not row.status:
            row.status = "new"

        # product: required
        if row.product_name and row.product_name.lower() not in products:
            valid = ", ".join(sorted(p.name for p in products.values()))
            errors.append(f"Unknown product '{row.product_name}' — valid: {valid}")
        elif not row.product_name:
            errors.append("Product is required")

        # guest_estimate: if raw value was non-empty but didn't parse to int
        if row.guest_estimate_raw and row.guest_estimate is None:
            errors.append(f"Invalid guest count '{row.guest_estimate_raw}' — must be a number")

        row.error = "; ".join(errors) if errors else ""


def load_xlsx(file_obj, sheet_name=None):
    """Load an Excel file. Returns (header, data_rows, sheet_names)."""
    import openpyxl

    wb = openpyxl.load_workbook(file_obj, read_only=True)
    sheet_names = wb.sheetnames

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], [], sheet_names

    header = rows[0]
    data_rows = rows[1:]
    return header, data_rows, sheet_names


def load_csv(file_obj):
    """Load a CSV file. Returns (header, data_rows, [])."""
    text = file_obj.read()
    if isinstance(text, bytes):
        text = text.decode("utf-8-sig")

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], [], []

    header = tuple(rows[0])
    data_rows = [tuple(r) for r in rows[1:]]
    return header, data_rows, []


def flag_duplicates(import_rows, org=None):
    """Mark rows whose email already exists in the DB. Warning only, does not block."""
    emails = [r.contact_email.lower() for r in import_rows if r.contact_email and not r.skipped]
    if not emails:
        return
    qs = Lead.objects.exclude(contact_email='')
    if org:
        qs = qs.filter(organisation=org)
    all_db_emails = set(qs.values_list("contact_email", flat=True))
    existing_lower = {e.lower() for e in all_db_emails}
    for row in import_rows:
        if row.contact_email and row.contact_email.lower() in existing_lower:
            row.duplicate_warning = True


def commit_rows(import_rows, org):
    """Create Lead objects for valid (non-skipped, non-error) rows. Returns (created_count, errors)."""
    from bookings.models.leads import ProductLine
    from users.models import User

    # Build lookup caches scoped to org
    product_names = {r.product_name.lower() for r in import_rows if r.product_name}
    user_emails = {r.assigned_to_email.lower() for r in import_rows if r.assigned_to_email}

    product_cache = {}
    if product_names:
        for p in ProductLine.objects.filter(organisation=org):
            if p.name.lower() in product_names:
                product_cache[p.name.lower()] = p

    user_cache = {}
    if user_emails:
        for u in User.objects.filter(organisation=org):
            if u.email.lower() in user_emails:
                user_cache[u.email.lower()] = u

    created = 0
    errors = []
    for row in import_rows:
        if row.skipped or row.error:
            continue

        row_product = product_cache.get(row.product_name.lower()) if row.product_name else None

        if not row_product:
            row.error = "Product line is required (set in CSV or select from dropdown)"
            errors.append(f"Row {row.row_num}: {row.error}")
            continue

        row_assigned = None
        if row.assigned_to_email:
            row_assigned = user_cache.get(row.assigned_to_email.lower())

        try:
            Lead.objects.create(
                organisation=org,
                contact_name=row.contact_name,
                contact_email=row.contact_email,
                contact_phone=row.contact_phone,
                event_type=row.event_type,
                guest_estimate=row.guest_estimate,
                event_date=row.event_date,
                lead_date=row.lead_date,
                source=row.source,
                status=row.status,
                notes=row.notes,
                product=row_product,
                assigned_to=row_assigned,
            )
            row.created = True
            created += 1
        except Exception as e:
            row.error = str(e)
            errors.append(f"Row {row.row_num}: {e}")
    return created, errors
